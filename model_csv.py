"""
Model (CSV): đọc/ghi dữ liệu điểm sinh viên bằng pandas.
Thống kê tính bằng numpy (vector hoá) + pandas (groupby).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd


REQUIRED_COLS = [
    "student_id",
    "full_name",
    "gender",
    "age",
    "course_code",
    "score",
    "credits",
    "notes",
]


class CsvModel:
    def __init__(self, csv_path: Path) -> None:
        self.csv_path = Path(csv_path)
        self.csv_path.parent.mkdir(parents=True, exist_ok=True)
        if not self.csv_path.exists():
            self._write_df(self._empty_df())

    def _empty_df(self) -> pd.DataFrame:
        return pd.DataFrame({c: pd.Series(dtype="object") for c in REQUIRED_COLS})

    def _read_df(self) -> pd.DataFrame:
        df = pd.read_csv(self.csv_path, encoding="utf-8-sig")
        for c in REQUIRED_COLS:
            if c not in df.columns:
                df[c] = "" if c != "age" else np.nan
        df = df[REQUIRED_COLS].copy()

        # Chuẩn hoá kiểu dữ liệu
        df["score"] = pd.to_numeric(df["score"], errors="coerce")
        df["credits"] = pd.to_numeric(df["credits"], errors="coerce")
        df["age"] = pd.to_numeric(df["age"], errors="coerce").astype("Int64")
        df["gender"] = df["gender"].fillna("").astype(str)
        df["notes"] = df["notes"].fillna("").astype(str)
        df["student_id"] = df["student_id"].fillna("").astype(str)
        df["full_name"] = df["full_name"].fillna("").astype(str)
        df["course_code"] = df["course_code"].fillna("").astype(str)
        return df

    def _write_df(self, df: pd.DataFrame) -> None:
        df.to_csv(self.csv_path, index=False, encoding="utf-8-sig")

    # --- CRUD (mỗi dòng = 1 môn của 1 SV) ---
    def list_rows(self) -> list[dict]:
        df = self._read_df()
        df = df.reset_index(drop=True)
        df.insert(0, "row_id", df.index + 1)  # ID hiển thị 1..N
        return df.to_dict(orient="records")

    def add_row(self, row: dict) -> None:
        df = self._read_df()
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        self._write_df(df[REQUIRED_COLS])

    def update_row(self, row_id: int, new_row: dict) -> None:
        df = self._read_df().reset_index(drop=True)
        idx = row_id - 1
        if idx < 0 or idx >= len(df):
            raise ValueError("Row ID không tồn tại.")
        for k in REQUIRED_COLS:
            df.at[idx, k] = new_row.get(k, df.at[idx, k])
        self._write_df(df[REQUIRED_COLS])

    def delete_rows(self, row_ids: list[int]) -> None:
        if not row_ids:
            return
        df = self._read_df().reset_index(drop=True)
        idxs = sorted({rid - 1 for rid in row_ids if rid > 0})
        mask = np.ones(len(df), dtype=bool)
        for i in idxs:
            if 0 <= i < len(df):
                mask[i] = False
        df2 = df.loc[mask].copy()
        self._write_df(df2[REQUIRED_COLS])

    # --- Import/Export ---
    def import_csv(self, other_csv: Path) -> int:
        other = pd.read_csv(other_csv, encoding="utf-8-sig")
        for c in REQUIRED_COLS:
            if c not in other.columns:
                other[c] = "" if c != "age" else np.nan
        other = other[REQUIRED_COLS].copy()
        df = self._read_df()
        df = pd.concat([df, other], ignore_index=True)
        self._write_df(df[REQUIRED_COLS])
        return len(other)

    def export_csv(self, out_path: Path) -> None:
        df = self._read_df()
        
        # Đổi tên cột sang tiếng Việt để xuất ra bảng đẹp và dễ hiểu hơn
        headers = {
            "student_id": "Mã Sinh Viên",
            "full_name": "Họ và Tên",
            "gender": "Giới Tính",
            "age": "Tuổi",
            "course_code": "Mã Môn Học",
            "score": "Điểm Số",
            "credits": "Số Tín Chỉ",
            "notes": "Ghi Chú"
        }
        df_export = df.rename(columns=headers)

        if out_path.suffix.lower() == ".xlsx":
            # Xuất file Excel (.xlsx) với định dạng bảng biểu cực kỳ đẹp và chuyên nghiệp
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_export.to_excel(writer, sheet_name="Danh sách điểm", index=False)
                
                workbook = writer.book
                worksheet = writer.sheets["Danh sách điểm"]
                
                # Bật hiển thị lưới (grid lines)
                worksheet.views.sheetView[0].showGridLines = True
                
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Định nghĩa font chữ & màu sắc nền tiêu đề (xanh Navy sang trọng)
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                
                # Dòng xen kẽ (Zebra rows) màu xám nhạt tinh tế
                body_font = Font(name="Segoe UI", size=10)
                zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
                
                # Căn lề
                align_center = Alignment(horizontal="center", vertical="center")
                align_left = Alignment(horizontal="left", vertical="center")
                
                # Đường viền mỏng màu xám
                thin_side = Side(border_style="thin", color="D9D9D9")
                border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                
                # Định dạng cho Header
                for col_num in range(1, len(df_export.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = border_all
                
                # Định dạng cho Body
                for row_idx in range(2, len(df_export) + 2):
                    is_zebra = (row_idx % 2 == 0)
                    for col_idx in range(1, len(df_export.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = body_font
                        cell.border = border_all
                        
                        if is_zebra:
                            cell.fill = zebra_fill
                        
                        col_name = df_export.columns[col_idx - 1]
                        if col_name in ["Mã Sinh Viên", "Giới Tính", "Tuổi", "Mã Môn Học", "Điểm Số", "Số Tín Chỉ"]:
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left
                            
                        # Định dạng kiểu hiển thị số
                        if col_name == "Điểm Số" and cell.value is not None:
                            try:
                                cell.value = float(cell.value)
                                cell.number_format = "0.00"
                            except ValueError:
                                pass
                        elif col_name == "Số Tín Chỉ" and cell.value is not None:
                            try:
                                cell.value = float(cell.value)
                                cell.number_format = "0.0"
                            except ValueError:
                                pass
                        elif col_name == "Tuổi" and cell.value is not None:
                            try:
                                cell.value = int(cell.value)
                                cell.number_format = "0"
                            except ValueError:
                                pass
                
                # Tự động điều chỉnh độ rộng cột vừa khít nội dung
                for col in worksheet.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            val_str = str(cell.value)
                            # Cân nhắc ký tự tiếng Việt chiếm nhiều không gian hơn
                            val_len = len(val_str.encode('utf-8')) - len(val_str) // 2
                            if val_len > max_len:
                                max_len = val_len
                    worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
        else:
            # Xuất ra file CSV tiêu chuẩn với encoding utf-8-sig để đọc tiếng Việt trên Excel không lỗi font
            df_export.to_csv(out_path, index=False, encoding="utf-8-sig")

    # --- Thống kê (NumPy + Pandas) ---
    def statistics(self) -> dict:
        df = self._read_df()
        if df.empty:
            return {
                "rows": 0,
                "students": 0,
                "mean_score": None,
                "mean_gpa": None,
                "gender": {},
                "notes_rows": 0,
                "total_credits": 0.0,
            }

        # Điểm trung bình tất cả dòng
        mean_score = float(np.nanmean(df["score"].to_numpy(dtype=np.float64)))
        total_credits = float(np.nansum(df["credits"].to_numpy(dtype=np.float64)))
        notes_rows = int(df["notes"].astype(str).str.strip().ne("").sum())

        # GPA theo SV: sum(score*credits)/sum(credits)
        gpas: list[float] = []
        for _, g in df.groupby("student_id"):
            sc = g["score"].to_numpy(dtype=np.float64)
            cr = g["credits"].to_numpy(dtype=np.float64)
            denom = np.nansum(cr)
            gpa = float(np.nansum(sc * cr) / denom) if denom > 0 else np.nan
            gpas.append(gpa)
        mean_gpa = float(np.nanmean(np.array(gpas))) if gpas else None

        students = int(df["student_id"].nunique())

        # Giới tính: lấy theo SV (dòng đầu tiên)
        gdf = df.copy()
        gdf["gender"] = gdf["gender"].fillna("").astype(str).replace("", "Chưa nhập")
        gdf = gdf.drop_duplicates("student_id")
        gender = gdf["gender"].value_counts().to_dict()

        return {
            "rows": int(len(df)),
            "students": students,
            "mean_score": mean_score,
            "mean_gpa": mean_gpa,
            "gender": gender,
            "notes_rows": notes_rows,
            "total_credits": total_credits,
        }

    def get_pie_chart_data(self) -> dict:
        """
        Phân loại điểm số (score) thành các nhóm học lực bằng Pandas và NumPy
        để phục vụ việc vẽ biểu đồ tròn ở phía View.
        """
        df = self._read_df()
        
        # Tạo sẵn một dict mặc định phòng trường hợp file CSV chưa có dữ liệu
        default_stats = {
            'Yếu/Kém (<5)': 0,
            'Trung bình/Khá (5->8)': 0,
            'Giỏi/Xuất sắc (>=8)': 0
        }
        
        if df.empty:
            return default_stats
            
        # Loại bỏ các dòng bị khuyết điểm (NaN) để không tính toán sai
        scores = df["score"].dropna().to_numpy()
        
        if len(scores) == 0:
            return default_stats

        # Sử dụng NumPy vectorization để phân loại nhóm điểm cực nhanh
        under_5 = int(np.sum(scores < 5))
        from_5_to_8 = int(np.sum((scores >= 5) & (scores < 8)))
        above_8 = int(np.sum(scores >= 8))

        return {
            'Yếu/Kém (<5)': under_5,
            'Trung bình/Khá (5->8)': from_5_to_8,
            'Giỏi/Xuất sắc (>=8)': above_8
        }

